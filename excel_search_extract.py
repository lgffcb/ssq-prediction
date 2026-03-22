#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 数据提取工具 - GUI 版本 v2
- 图形界面，支持拖拽和粘贴文件路径
- 自动查找 sheet 名包含"明细"的工作表（支持"分包明细表"等）
- 显示所有隐藏表格、取消行列隐藏、公式转数值
- 只提取用户指定的列
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
import os
import threading

class ExcelExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 数据提取工具")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        self.file_path = tk.StringVar()
        self.search_term = tk.StringVar()
        self.status_text = tk.StringVar()
        self.status_text.set("就绪")
        
        # 存储列信息
        self.columns_list = []
        self.column_vars = {}
        
        self.create_widgets()
    
    def create_widgets(self):
        """创建界面组件"""
        
        # === 标题 ===
        title_frame = ttk.Frame(self.root, padding="10")
        title_frame.pack(fill=tk.X)
        
        title_label = ttk.Label(
            title_frame,
            text="📊 Excel 数据提取工具",
            font=("Microsoft YaHei UI", 16, "bold")
        )
        title_label.pack()
        
        subtitle_label = ttk.Label(
            title_frame,
            text="自动提取明细表数据 - 支持隐藏表格显示和公式转数值",
            font=("Microsoft YaHei UI", 9)
        )
        subtitle_label.pack()
        
        # === 文件选择区域 ===
        file_frame = ttk.LabelFrame(self.root, text="1️⃣ 选择 Excel 文件", padding="10")
        file_frame.pack(fill=tk.X, padx=20, pady=10)
        
        file_entry_frame = ttk.Frame(file_frame)
        file_entry_frame.pack(fill=tk.X)
        
        self.file_entry = ttk.Entry(
            file_entry_frame,
            textvariable=self.file_path,
            font=("Consolas", 10)
        )
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.file_entry.bind('<Control-v>', self.on_paste)
        self.file_entry.bind('<Control-V>', self.on_paste)
        
        browse_btn = ttk.Button(
            file_entry_frame,
            text="📂 浏览...",
            command=self.browse_file,
            width=10
        )
        browse_btn.pack(side=tk.RIGHT)
        
        # === 查找内容区域 ===
        search_frame = ttk.LabelFrame(self.root, text="2️⃣ 输入查找内容", padding="10")
        search_frame.pack(fill=tk.X, padx=20, pady=10)
        
        search_entry_frame = ttk.Frame(search_frame)
        search_entry_frame.pack(fill=tk.X)
        
        self.search_entry = ttk.Entry(
            search_entry_frame,
            textvariable=self.search_term,
            font=("Consolas", 11),
            width=50
        )
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.search_entry.bind('<Return>', lambda e: self.load_columns())
        
        hint_label = ttk.Label(
            search_frame,
            text="💡 提示：输入要查找的内容，按回车加载列选项",
            foreground="gray"
        )
        hint_label.pack(anchor=tk.W, pady=(5, 0))
        
        # === 选择列区域 ===
        columns_frame = ttk.LabelFrame(self.root, text="3️⃣ 选择要提取的列（勾选需要的列）", padding="10")
        columns_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # 列选择框架（带滚动条）
        columns_canvas_frame = ttk.Frame(columns_frame)
        columns_canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        self.columns_canvas = tk.Canvas(columns_canvas_frame, height=150)
        columns_scrollbar = ttk.Scrollbar(columns_canvas_frame, orient="vertical", command=self.columns_canvas.yview)
        self.columns_inner_frame = ttk.Frame(self.columns_canvas)
        
        self.columns_inner_frame.bind(
            "<Configure>",
            lambda e: self.columns_canvas.configure(scrollregion=self.columns_canvas.bbox("all"))
        )
        
        self.columns_canvas.create_window((0, 0), window=self.columns_inner_frame, anchor="nw")
        self.columns_canvas.configure(yscrollcommand=columns_scrollbar.set)
        
        self.columns_canvas.pack(side="left", fill="both", expand=True)
        columns_scrollbar.pack(side="right", fill="y")
        
        # 全选/反选按钮
        select_btn_frame = ttk.Frame(columns_frame)
        select_btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        select_all_btn = ttk.Button(
            select_btn_frame,
            text="✓ 全选",
            command=self.select_all_columns,
            width=10
        )
        select_all_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        deselect_all_btn = ttk.Button(
            select_btn_frame,
            text="✗ 取消全选",
            command=self.deselect_all_columns,
            width=10
        )
        deselect_all_btn.pack(side=tk.LEFT, padx=5)
        
        # === 开始按钮 ===
        btn_frame = ttk.Frame(self.root, padding="10")
        btn_frame.pack(fill=tk.X, padx=20, pady=10)
        
        self.load_btn = ttk.Button(
            btn_frame,
            text="📋 加载列选项",
            command=self.load_columns,
            width=20
        )
        self.load_btn.pack(pady=5)
        
        self.start_btn = ttk.Button(
            btn_frame,
            text="🚀 开始提取",
            command=self.start_extraction,
            width=20
        )
        self.start_btn.pack(pady=5)
        
        # === 进度/状态区域 ===
        progress_frame = ttk.LabelFrame(self.root, text="📋 处理状态", padding="10")
        progress_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        status_bar = ttk.Label(
            progress_frame,
            textvariable=self.status_text,
            font=("Consolas", 9),
            foreground="blue"
        )
        status_bar.pack(anchor=tk.W, pady=(0, 10))
        
        self.log_text = tk.Text(
            progress_frame,
            height=10,
            font=("Consolas", 9),
            wrap=tk.WORD,
            state=tk.DISABLED
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # === 底部按钮 ===
        bottom_frame = ttk.Frame(self.root, padding="10")
        bottom_frame.pack(fill=tk.X, padx=20, pady=10)
        
        open_folder_btn = ttk.Button(
            bottom_frame,
            text="📁 打开结果文件夹",
            command=self.open_result_folder,
            state=tk.DISABLED
        )
        open_folder_btn.pack(side=tk.LEFT)
        self.open_folder_btn = open_folder_btn
        
        copyright_label = ttk.Label(
            bottom_frame,
            text="Excel 数据提取工具 v2.0",
            foreground="gray",
            font=("Microsoft YaHei UI", 8)
        )
        copyright_label.pack(side=tk.RIGHT)
        
        self.result_file_path = None
    
    def on_paste(self, event):
        """处理粘贴事件"""
        try:
            clipboard = self.root.clipboard_get()
            if os.path.exists(clipboard) and clipboard.endswith(('.xlsx', '.xlsm')):
                self.file_path.set(clipboard)
                self.log("✅ 已粘贴文件路径")
                return 'break'
            return None
        except:
            return None
    
    def browse_file(self):
        """浏览选择文件"""
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.file_path.set(file_path)
            self.log(f"✅ 已选择：{os.path.basename(file_path)}")
    
    def log(self, message):
        """添加日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
    
    def clear_log(self):
        """清空日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
    
    def select_all_columns(self):
        """全选所有列"""
        for var in self.column_vars.values():
            var.set(True)
    
    def deselect_all_columns(self):
        """取消全选"""
        for var in self.column_vars.values():
            var.set(False)
    
    def load_columns(self):
        """加载列选项"""
        if not self.file_path.get():
            messagebox.showerror("错误", "请先选择 Excel 文件！")
            return
        
        if not os.path.exists(self.file_path.get()):
            messagebox.showerror("错误", "文件不存在！")
            return
        
        if not self.search_term.get():
            messagebox.showerror("错误", "请输入查找内容！")
            return
        
        self.load_btn.config(state=tk.DISABLED)
        self.clear_log()
        
        thread = threading.Thread(target=self._load_columns_thread)
        thread.daemon = True
        thread.start()
    
    def _load_columns_thread(self):
        """加载列选项线程"""
        try:
            file_path = self.file_path.get()
            
            self.log("🔍 正在分析 Excel 文件...")
            self.root.update()
            
            # 读取 Excel
            all_sheets = pd.read_excel(file_path, sheet_name=None)
            sheet_names = list(all_sheets.keys())
            
            self.log(f"📊 文件包含 {len(sheet_names)} 个工作表")
            
            # 查找明细表（支持"分包明细表"、"明细表"等）
            target_sheet = None
            for sheet_name in sheet_names:
                if '明细' in sheet_name:
                    target_sheet = sheet_name
                    break
            
            if not target_sheet:
                self.log(f"❌ 未找到包含'明细'的工作表")
                self.log(f"   可用的工作表：{', '.join(sheet_names)}")
                messagebox.showerror("错误", f"未找到包含'明细'的工作表\n\n可用的工作表：\n{', '.join(sheet_names)}")
                self.load_btn.config(state=tk.NORMAL)
                return
            
            self.log(f"✅ 找到明细工作表：【{target_sheet}】")
            
            df = all_sheets[target_sheet]
            self.log(f"工作表大小：{len(df)} 行 × {len(df.columns)} 列")
            
            # 存储列信息
            self.columns_list = []
            for i, col in enumerate(df.columns):
                col_letter = self.get_column_letter(i)
                self.columns_list.append((i, col_letter, col))
            
            # 在 UI 中显示列选项
            self.root.after(0, self._update_columns_ui)
            
            self.log("✅ 列选项已加载，请勾选需要提取的列")
            
        except Exception as e:
            self.log(f"❌ 错误：{str(e)}")
            messagebox.showerror("错误", f"加载失败：\n{str(e)}")
        
        finally:
            self.root.after(0, lambda: self.load_btn.config(state=tk.NORMAL))
    
    def _update_columns_ui(self):
        """更新列选项 UI"""
        # 清空现有选项
        for widget in self.columns_inner_frame.winfo_children():
            widget.destroy()
        
        self.column_vars = {}
        
        # 创建复选框（每行 5 个）
        cols_per_row = 5
        for i, (col_idx, col_letter, col_name) in enumerate(self.columns_list):
            row = i // cols_per_row
            col = i % cols_per_row
            
            var = tk.BooleanVar(value=False)
            self.column_vars[col_idx] = var
            
            cb = ttk.Checkbutton(
                self.columns_inner_frame,
                text=f"{col_letter}列：{col_name}",
                variable=var,
                onvalue=True,
                offvalue=False
            )
            cb.grid(row=row, column=col, sticky=tk.W, padx=5, pady=3)
        
        # 自动选择包含查找内容的列
        for col_idx, col_letter, col_name in self.columns_list:
            if self.search_term.get() in str(col_name):
                self.column_vars[col_idx].set(True)
    
    def start_extraction(self):
        """开始提取"""
        if not self.file_path.get():
            messagebox.showerror("错误", "请选择 Excel 文件！")
            return
        
        if not os.path.exists(self.file_path.get()):
            messagebox.showerror("错误", "文件不存在！")
            return
        
        if not self.search_term.get():
            messagebox.showerror("错误", "请输入查找内容！")
            return
        
        # 检查是否选择了列
        selected_cols = [idx for idx, var in self.column_vars.items() if var.get()]
        if not selected_cols:
            messagebox.showerror("错误", "请至少选择一列！")
            return
        
        self.start_btn.config(state=tk.DISABLED)
        self.clear_log()
        
        thread = threading.Thread(target=self.run_extraction)
        thread.daemon = True
        thread.start()
    
    def run_extraction(self):
        """运行提取逻辑"""
        try:
            file_path = self.file_path.get()
            search_term = self.search_term.get()
            selected_cols = [idx for idx, var in self.column_vars.items() if var.get()]
            
            self.log("="*60)
            self.log("🔧 正在处理 Excel 文件...")
            self.root.update()
            
            # 处理 Excel
            processed_file = self.unhide_and_convert_values(file_path)
            if not processed_file:
                self.log("❌ Excel 文件处理失败")
                self.start_btn.config(state=tk.NORMAL)
                return
            
            self.log("✅ Excel 文件处理完成")
            self.root.update()
            
            # 搜索并提取
            self.log("🔍 正在搜索并提取数据...")
            self.root.update()
            
            result = self.search_and_extract(processed_file, search_term, selected_cols)
            
            if result is not None:
                self.log("="*60)
                self.log("✅ 数据提取完成！")
                self.log(f"📁 结果已保存到：{self.result_file_path}")
                self.open_folder_btn.config(state=tk.NORMAL)
                messagebox.showinfo("成功", f"数据提取完成！\n结果已保存到：\n{self.result_file_path}")
            else:
                self.log("❌ 未找到匹配的数据")
                messagebox.showwarning("提示", "未找到匹配的数据")
            
        except Exception as e:
            self.log(f"❌ 错误：{str(e)}")
            messagebox.showerror("错误", f"提取失败：\n{str(e)}")
        
        finally:
            self.start_btn.config(state=tk.NORMAL)
    
    def unhide_and_convert_values(self, file_path):
        """处理 Excel 文件"""
        try:
            wb = load_workbook(file_path)
            
            # 显示所有隐藏的工作表
            hidden_sheets = []
            for sheet in wb.worksheets:
                if sheet.sheet_state == 'hidden':
                    sheet.sheet_state = 'visible'
                    hidden_sheets.append(sheet.title)
            
            if hidden_sheets:
                self.log(f"   ✅ 显示了 {len(hidden_sheets)} 个隐藏的工作表：{hidden_sheets}")
            
            # 取消行列隐藏
            total_hidden_rows = 0
            total_hidden_cols = 0
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                for row_dim in sheet.row_dimensions.values():
                    if row_dim.hidden:
                        row_dim.hidden = False
                        total_hidden_rows += 1
                
                for col_dim in sheet.column_dimensions.values():
                    if col_dim.hidden:
                        col_dim.hidden = False
                        total_hidden_cols += 1
            
            self.log(f"   ✅ 取消了 {total_hidden_rows} 个隐藏行")
            self.log(f"   ✅ 取消了 {total_hidden_cols} 个隐藏列")
            
            # 保存为临时文件（公式转数值）
            temp_file = os.path.splitext(file_path)[0] + '_processed.xlsx'
            
            from openpyxl import Workbook
            wb_new = Workbook()
            wb_new.remove(wb_new.active)
            
            for sheet_name in wb.sheetnames:
                old_sheet = wb[sheet_name]
                new_sheet = wb_new.create_sheet(title=sheet_name)
                
                for row in old_sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        new_sheet.append(row)
                
                for col_letter, col_dim in old_sheet.column_dimensions.items():
                    if col_dim.width:
                        new_sheet.column_dimensions[col_letter].width = col_dim.width
            
            wb_new.save(temp_file)
            self.log(f"   ✅ 公式已转换为数值")
            
            return temp_file
            
        except Exception as e:
            self.log(f"   ❌ 处理失败：{str(e)}")
            return None
    
    def get_column_letter(self, col_index):
        """将列索引转换为 Excel 列字母"""
        result = ""
        while col_index >= 0:
            result = chr(ord('A') + (col_index % 26)) + result
            col_index = col_index // 26 - 1
        return result
    
    def search_and_extract(self, file_path, search_term, selected_cols):
        """搜索并提取指定的列"""
        try:
            all_sheets = pd.read_excel(file_path, sheet_name=None)
            sheet_names = list(all_sheets.keys())
            
            self.log(f"📊 文件包含 {len(sheet_names)} 个工作表")
            
            # 查找明细表
            target_sheet = None
            for sheet_name in sheet_names:
                if '明细' in sheet_name:
                    target_sheet = sheet_name
                    break
            
            if not target_sheet:
                self.log(f"❌ 未找到包含'明细'的工作表")
                return None
            
            self.log(f"✅ 找到明细工作表：【{target_sheet}】")
            
            df = all_sheets[target_sheet]
            self.log(f"工作表大小：{len(df)} 行 × {len(df.columns)} 列")
            
            # 将搜索列转换为字符串以便搜索
            df_str = df.astype(str)
            
            # 搜索包含查找内容的行
            mask = df_str.apply(lambda col: col.str.contains(str(search_term), na=False, case=False)).any(axis=1)
            matched_rows = df[mask]
            
            if len(matched_rows) == 0:
                self.log(f"❌ 未找到包含'{search_term}'的行")
                return None
            
            self.log(f"✅ 找到 {len(matched_rows)} 条匹配记录")
            
            # 提取用户选择的列
            result = matched_rows.iloc[:, selected_cols].copy()
            
            # 重命名列，添加列字母
            new_columns = []
            for col_idx in selected_cols:
                col_letter = self.get_column_letter(col_idx)
                col_name = df.columns[col_idx]
                new_columns.append(f'{col_letter}列：{col_name}')
            
            result.columns = new_columns
            
            # 保存结果
            self.result_file_path = os.path.splitext(file_path)[0] + '_提取结果.xlsx'
            result.to_excel(self.result_file_path, index=False)
            
            self.log(f"\n📋 提取结果:")
            self.log(f"查找内容：{search_term}")
            self.log(f"匹配行数：{len(result)}")
            self.log(f"提取列：{', '.join(result.columns)}")
            
            return result
            
        except Exception as e:
            self.log(f"❌ 错误：{str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def open_result_folder(self):
        """打开结果文件夹"""
        if self.result_file_path and os.path.exists(self.result_file_path):
            folder_path = os.path.dirname(self.result_file_path)
            os.startfile(folder_path)

def main():
    root = tk.Tk()
    app = ExcelExtractorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
