#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抖音视频批量数据汇总 Excel
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 3 个视频的数据
videos = [
    {
        "视频 ID": "7616988650777087295",
        "标题": "情绪价值拉满的语言技巧，推拉技巧",
        "作者": "灯哥说脱单",
        "作者主页": "https://www.douyin.com/user/MS4wLjABAAAAOAwjxUrLpBR7ZZMnEHH_c13TCmAeyH8J3DUTPeg8LHV8egfwPIPveH9mEJGHY_GO",
        "粉丝数": 9385,
        "总获赞": "1.7 万",
        "发布时间": "2026-03-14 19:27",
        "视频时长": "04:22",
        "点赞数": 2320,
        "评论数": 55,
        "收藏数": 1016,
        "分享数": 362,
        "话题标签": "#干货分享 #推拉技巧 #情绪价值 #恋爱心理学 #恋爱",
        "视频链接": "https://v.douyin.com/KE4eK_vzT0Y/"
    },
    {
        "视频 ID": "7619243018100592987",
        "标题": "第 64 集 | 防御性驾驶技巧学起来！",
        "作者": "抚州交警",
        "作者主页": "https://www.douyin.com/user/MS4wLjABAAAApNeEgUHaNZ-bvqcXWFsGtqreWMNbQGku2sYT8bJgxEA",
        "粉丝数": 600000,
        "总获赞": "1743.8 万",
        "发布时间": "2026-03-21 07:50",
        "视频时长": "00:45",
        "点赞数": 46000,
        "评论数": 1283,
        "收藏数": 22000,
        "分享数": 26000,
        "话题标签": "#知识科普 #防御性驾驶 #政务原创作者联盟 #开车技巧",
        "视频链接": "https://v.douyin.com/oCbxegyfZhk/"
    },
    {
        "视频 ID": "7619264361403534634",
        "标题": "使用军用 AI 打击我永兴岛？美国 AI 军火商露出獠牙（上）",
        "作者": "域与局",
        "作者主页": "https://www.douyin.com/user/MS4wLjABAAAAm60apUhYL4jFCHVbvAYnrlKvfvGSyt0FEcfhXv-6X_0hMkrc1r4zpSsqP7nZRTzU",
        "粉丝数": 585000,
        "总获赞": "217.7 万",
        "发布时间": "2026-03-20 19:20",
        "视频时长": "06:16",
        "点赞数": 21000,
        "评论数": 1153,
        "收藏数": 7137,
        "分享数": 3651,
        "话题标签": "#时代的荣耀 #全球创作者计划 #零基础看懂全球 #硬核深度计划",
        "视频链接": "https://v.douyin.com/cR3QGNmWJYc/"
    }
]

# 创建 Excel 工作簿
wb = Workbook()

# Sheet 1: 汇总数据
ws_summary = wb.active
ws_summary.title = "视频汇总"

# 定义样式
title_font = Font(bold=True, size=14, color="000000")
header_font = Font(bold=True, size=11, color="FFFFFF")
normal_font = Font(size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
blue_font = Font(color="0000FF")
green_font = Font(color="008000")

# 表头
headers = ["序号", "视频标题", "作者", "粉丝数", "发布时间", "视频时长", "点赞数", "评论数", "收藏数", "分享数", "总互动量", "互动率", "话题标签", "视频链接"]

for col, header in enumerate(headers, start=1):
    cell = ws_summary[f'{get_column_letter(col)}1']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 数据行
for idx, video in enumerate(videos, start=2):
    total_engagement = video["点赞数"] + video["评论数"] + video["收藏数"] + video["分享数"]
    engagement_rate = (total_engagement / video["粉丝数"]) * 100 if video["粉丝数"] > 0 else 0
    
    ws_summary[f'A{idx}'] = idx - 1
    ws_summary[f'B{idx}'] = video["标题"]
    ws_summary[f'C{idx}'] = video["作者"]
    ws_summary[f'D{idx}'] = video["粉丝数"]
    ws_summary[f'E{idx}'] = video["发布时间"]
    ws_summary[f'F{idx}'] = video["视频时长"]
    ws_summary[f'G{idx}'] = video["点赞数"]
    ws_summary[f'H{idx}'] = video["评论数"]
    ws_summary[f'I{idx}'] = video["收藏数"]
    ws_summary[f'J{idx}'] = video["分享数"]
    ws_summary[f'K{idx}'] = total_engagement  # 总互动量
    ws_summary[f'L{idx}'] = engagement_rate / 100  # 互动率（百分比格式）
    ws_summary[f'M{idx}'] = video["话题标签"]
    ws_summary[f'N{idx}'] = video["视频链接"]
    
    # 设置格式
    for col in range(1, 15):
        cell = ws_summary[f'{get_column_letter(col)}{idx}']
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 链接用绿色
    ws_summary[f'N{idx}'].font = green_font
    # 互动率用百分比格式
    ws_summary[f'L{idx}'].number_format = "0.00%"

# 设置列宽
ws_summary.column_dimensions['A'].width = 6
ws_summary.column_dimensions['B'].width = 45
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 12
ws_summary.column_dimensions['E'].width = 16
ws_summary.column_dimensions['F'].width = 10
ws_summary.column_dimensions['G'].width = 10
ws_summary.column_dimensions['H'].width = 10
ws_summary.column_dimensions['I'].width = 10
ws_summary.column_dimensions['J'].width = 10
ws_summary.column_dimensions['K'].width = 12
ws_summary.column_dimensions['L'].width = 10
ws_summary.column_dimensions['M'].width = 35
ws_summary.column_dimensions['N'].width = 50

# Sheet 2: 详细数据
ws_detail = wb.create_sheet("详细信息")

detail_headers = ["字段", "视频 1", "视频 2", "视频 3"]
for col, header in enumerate(detail_headers, start=1):
    cell = ws_detail[f'{get_column_letter(col)}1']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

detail_fields = [
    ("视频 ID", "视频 ID", "视频 ID", "视频 ID"),
    ("标题", videos[0]["标题"], videos[1]["标题"], videos[2]["标题"]),
    ("作者", videos[0]["作者"], videos[1]["作者"], videos[2]["作者"]),
    ("作者主页", videos[0]["作者主页"], videos[1]["作者主页"], videos[2]["作者主页"]),
    ("粉丝数", videos[0]["粉丝数"], videos[1]["粉丝数"], videos[2]["粉丝数"]),
    ("总获赞", videos[0]["总获赞"], videos[1]["总获赞"], videos[2]["总获赞"]),
    ("发布时间", videos[0]["发布时间"], videos[1]["发布时间"], videos[2]["发布时间"]),
    ("视频时长", videos[0]["视频时长"], videos[1]["视频时长"], videos[2]["视频时长"]),
    ("点赞数", videos[0]["点赞数"], videos[1]["点赞数"], videos[2]["点赞数"]),
    ("评论数", videos[0]["评论数"], videos[1]["评论数"], videos[2]["评论数"]),
    ("收藏数", videos[0]["收藏数"], videos[1]["收藏数"], videos[2]["收藏数"]),
    ("分享数", videos[0]["分享数"], videos[1]["分享数"], videos[2]["分享数"]),
    ("话题标签", videos[0]["话题标签"], videos[1]["话题标签"], videos[2]["话题标签"]),
    ("视频链接", videos[0]["视频链接"], videos[1]["视频链接"], videos[2]["视频链接"]),
]

for row_idx, (field, v1, v2, v3) in enumerate(detail_fields, start=2):
    ws_detail[f'A{row_idx}'] = field
    ws_detail[f'B{row_idx}'] = v1
    ws_detail[f'C{row_idx}'] = v2
    ws_detail[f'D{row_idx}'] = v3
    
    for col in range(1, 5):
        cell = ws_detail[f'{get_column_letter(col)}{row_idx}']
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 链接用绿色
    if field == "视频链接" or field == "作者主页":
        for col in range(2, 5):
            ws_detail[f'{get_column_letter(col)}{row_idx}'].font = green_font

ws_detail.column_dimensions['A'].width = 15
ws_detail.column_dimensions['B'].width = 45
ws_detail.column_dimensions['C'].width = 45
ws_detail.column_dimensions['D'].width = 45

# 添加统计摘要
summary_row = len(detail_fields) + 3
ws_detail[f'A{summary_row}'] = "数据统计摘要"
ws_detail[f'A{summary_row}'].font = title_font

summary_row += 1
total_likes = sum(v["点赞数"] for v in videos)
total_comments = sum(v["评论数"] for v in videos)
total_favorites = sum(v["收藏数"] for v in videos)
total_shares = sum(v["分享数"] for v in videos)
total_all = total_likes + total_comments + total_favorites + total_shares

ws_detail[f'A{summary_row}'] = "总点赞数"
ws_detail[f'B{summary_row}'] = total_likes
ws_detail[f'A{summary_row+1}'] = "总评论数"
ws_detail[f'B{summary_row+1}'] = total_comments
ws_detail[f'A{summary_row+2}'] = "总收藏数"
ws_detail[f'B{summary_row+2}'] = total_favorites
ws_detail[f'A{summary_row+3}'] = "总分享数"
ws_detail[f'B{summary_row+3}'] = total_shares
ws_detail[f'A{summary_row+4}'] = "总互动量"
ws_detail[f'B{summary_row+4}'] = total_all
ws_detail[f'A{summary_row+5}'] = "平均互动率"
avg_engagement_rate = sum((v["点赞数"]+v["评论数"]+v["收藏数"]+v["分享数"])/v["粉丝数"] for v in videos) / 3
ws_detail[f'B{summary_row+5}'] = avg_engagement_rate
ws_detail[f'B{summary_row+5}'].number_format = "0.00%"

# 添加抓取时间
ws_detail[f'A{summary_row+7}'] = "抓取时间"
ws_detail[f'B{summary_row+7}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# 保存文件
output_file = "/home/admin/openclaw/workspace/output/抖音视频批量汇总_3 个视频.xlsx"
wb.save(output_file)

print(f"✅ Excel 文件已创建：{output_file}")
print(f"\n📊 汇总统计：")
print(f"   视频总数：{len(videos)}")
print(f"   总点赞数：{total_likes:,}")
print(f"   总评论数：{total_comments:,}")
print(f"   总收藏数：{total_favorites:,}")
print(f"   总分享数：{total_shares:,}")
print(f"   总互动量：{total_all:,}")
print(f"   平均互动率：{avg_engagement_rate*100:.2f}%")
