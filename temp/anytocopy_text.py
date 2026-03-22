#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AnyToCopy 提取的完整文案保存到 Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# 完整文案
video_text = """我可以亲你吗？想的话就不对你说哈喽，大家好，我是登登，今天我们来讲推拉技巧。我们生活中很多地方都可以看到这个技巧的影子，比如综艺节目还有平常朋友聊天打趣。不过，本期视频里说的推拉，主要是针对恋爱里面的推拉。推拉技巧，简单来说，就是在相处过程中一进一退的方式，让对方产生大幅度的情绪波动，让对方情绪出现反差，犹如过山车一般，以此来达到情绪升温、拉近彼此关系、营造浪漫情绪体验的效果。 比如刚才这个视频，男生首先问"可以亲你吗？"，在得到了女生的答复以后，却说是骗他的，只是推；然后在女生失望的时候，又突然亲了上去，这是拉。瞬间就把刚才拉低的预期一下子拉到了最高点，效果直接拉满。大家可以体会一下，如果没有前面的推，那样效果肯定是要差很多的。而有了这个推拉的过程，通过先把对方的预期拉低，然后再突然拉高，就形成了一个过山车式的情绪波动，效果会远高于正常的预期，彼此的关系也会越来越近。 我们再来看一个推拉技巧的例子。这个视频同样是先推开，通过"推"来拉低预期，制造情绪的低谷，然后再通过一句情话将预期拉到高点。前面两个例子都是先推后拉，也有一些地方可以使用先拉后推。先推后拉通常是在说情话或者撩拨对方的时候，能让后面的撩拨威力大大增加；而先拉后推，则更多的是用于维持自身框架的情况。 我们也来看一个例子：如果我跟林志玲掉到海里，你要救谁？小 S 首先抛出送命题："当然是应该先救你，因为林志玲的个子比较高，她是可以自己露出水面来的。"因为那个高度，已经感觉你嫌我矮。"没有没有，你才矮，你给我出来站，你给我站出来！但是你比例好，你知道吗？"黄渤先是迎合小 S 说先救她，这里是"拉"；然后又通过说"林志玲高"来调侃小 S 矮，这里是"推"；然后又说"小 S 的比例好"，这里又是"拉"。三句话就让小 S 情绪像过山车一样，情绪价值直接拉满，关键是自己的框架也没有丢失。如果不知道框架是什么的，可以进我主页看我关于框架的那一期视频。 另外，推拉技巧不仅仅是语言上的，也可以搭配行为一起使用。比如上一个系列更新的"扛影撩拨"，许子权的画面，身体往男生方向前倾，并配合暧昧的语言："聊不聊男生？"这里是行为上和语言上的"拉"。就在男生有想亲吻的想法时，立马后退说："晚安，再见。"这里是"推"。这一段就是行为和语言上的同步推拉。 当然也有单纯行为上的推拉，比如下面这个片段：这里女生是在帮男生修指甲，男生若是想看指甲修得怎么样了，通常会把手抽回来看，而不是把脸贴过去。所以，男生这个行为更多的是在做一个测试，看女生的接受程度。距离近了，自然就会产生暧昧的感觉。像现实生活中，你们也可以通过这样的方式去测试女生对你的好感度，比如帮她系安全带的时候，盯着她看一小会儿，看她的反应。 好，继续回到主题。男生在看到女生有接吻意愿后，反而抽身拉开了距离，这里则是"推"。这里很多人可能想不通，为什么女生都有接吻意愿了，却又放弃这个窗口，然后拉远距离呢？这里就不得不说到一种男生了，网络上大家都叫他们"阿尔法男"。他们更喜欢去主导感情，而不是让女生去主导。这一类男生有很强大的自律能力和掌控能力，这类男人往往对异性有更大的吸引力。 看到这里应该知道什么是推拉技巧了吧？如果还不懂的，可以进我主页看看其他视频。好了，这一期就到这里了，喜欢登登的视频别忘了点赞关注，我们下期再见，期待！"""

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = "完整文案"

# 样式
title_font = Font(bold=True, size=14)
header_font = Font(bold=True, size=11, color="FFFFFF")
normal_font = Font(size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# 基本信息
ws['A1'] = "视频文案提取结果"
ws['A1'].font = title_font

ws['A3'] = "视频链接"
ws['B3'] = "https://v.douyin.com/KE4eK_vzT0Y/"

ws['A4'] = "视频标题"
ws['B4'] = "情绪价值拉满的语言技巧，推拉技巧"

ws['A5'] = "话题标签"
ws['B5'] = "#干货分享 #推拉技巧 #情绪价值 #恋爱心理学 #恋爱"

ws['A6'] = "提取时间"
ws['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

ws['A8'] = "完整文案内容"
ws['A8'].font = Font(bold=True, size=12)

ws['A9'] = video_text

# 格式设置
for col in ['A', 'B']:
    ws.column_dimensions[col].width = 80

ws['A9'].alignment = Alignment(wrap_text=True)

# 按段落拆分
ws2 = wb.create_sheet("分段文案")

paragraphs = video_text.split('。 ')
for idx, para in enumerate(paragraphs, start=1):
    ws2[f'A{idx}'] = f"段落{idx}"
    ws2[f'B{idx}'] = para + "。" if not para.endswith('。') else para

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 100

# 按内容结构拆分
ws3 = wb.create_sheet("内容结构")

structure = [
    ("开场白", "哈喽，大家好，我是登登，今天我们来讲推拉技巧"),
    ("定义", "推拉技巧，简单来说，就是在相处过程中一进一退的方式，让对方产生大幅度的情绪波动"),
    ("例子 1", "男生问可以亲你吗，先推后拉，情绪过山车"),
    ("例子 2", "黄渤和小 S 的对话，三句话情绪波动，拉推拉"),
    ("行为推拉", "身体前倾是拉，后退是推，行为和语言同步"),
    ("测试好感", "修指甲时把脸贴过去测试接受程度"),
    ("阿尔法男", "主导感情，自律和掌控能力，对异性更有吸引力"),
    ("结尾", "点赞关注，下期再见")
]

headers = ["结构", "内容摘要"]
for col, header in enumerate(headers, start=1):
    cell = ws3[f'{get_column_letter(col)}1']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill

for row, (struct, content) in enumerate(structure, start=2):
    ws3[f'A{row}'] = struct
    ws3[f'B{row}'] = content

ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 80

# 保存
output_file = "/home/admin/openclaw/workspace/output/抖音视频完整文案_推拉技巧.xlsx"
wb.save(output_file)

print(f"✅ 文案 Excel 已创建：{output_file}")
print(f"\n📝 文案统计：")
print(f"   总字数：{len(video_text)}")
print(f"   段落数：{len(paragraphs)}")
print(f"   结构数：{len(structure)}")
