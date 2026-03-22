#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抖音视频文案提取汇总
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 3 个视频的文案数据
videos = [
    {
        "视频 ID": "7616988650777087295",
        "标题": "情绪价值拉满的语言技巧，推拉技巧",
        "作者": "灯哥说脱单",
        "视频链接": "https://v.douyin.com/KE4eK_vzT0Y/",
        "视频时长": "04:22",
        "发布时间": "2026-03-14 19:27",
        "简介文案": "情绪价值拉满的语言技巧，推拉技巧 #干货分享 #推拉技巧 #情绪价值 #恋爱心理学 #恋爱",
        "作者声明": "虚构演绎，仅供娱乐",
        "章节要点": """【总述】
恋爱中的推拉技巧，通过一进一退的方式让对方情绪波动，拉近彼此关系。先推后拉和先拉后推都有不同的应用场景，可搭配语言和行为使用。掌握推拉技巧有助于营造浪漫情绪体验，提高吸引力。

【章节 1 - 引言】00:12
- 引入主题

【章节 2 - 推拉技巧的定义】00:21
- 推拉技巧是在相处中一进一退，让对方情绪波动，拉近关系

【章节 3 - 先推后拉】00:39
- 通过先拉低预期再突然拉高，形成情绪波动，效果更佳

【章节 4 - 先拉后推】01:30
- 用于维持自身框架，增加撩拨威力

【章节 5 - 行为上的推拉】02:23
- 推拉技巧可搭配行为使用，如身体前倾和后退，测试女生好感度

【章节 6 - 结语】04:11
- 总结""",
        "弹幕/热评": [
            "推拉技巧牛逼啊",
            "这就是大神的思路吗",
            "我看过全部视频…判刑了…",
            "我是一分钟都不敢快进啊，单身怕了",
            "全体起立！判决如下－死刑！"
        ]
    },
    {
        "视频 ID": "7619243018100592987",
        "标题": "第 64 集 | 防御性驾驶技巧学起来！",
        "作者": "抚州交警",
        "视频链接": "https://v.douyin.com/oCbxegyfZhk/",
        "视频时长": "00:45",
        "发布时间": "2026-03-21 07:50",
        "简介文案": "第 64 集 | 防御性驾驶技巧学起来！#知识科普 #防御性驾驶 #政务原创作者联盟 #开车技巧",
        "作者声明": "",
        "章节要点": """【视频主题】
防御性驾驶技巧教学

【核心内容】
- 会车时看车尾的重要性
- 右转时注意后方来车
- 左转时观察左右两侧
- 非机动车过路口应减速

【作者信息】
抚州交警 - 粉丝 60.0 万 获赞 1743.8 万
政务原创作者联盟成员""",
        "弹幕/热评": [
            "这才是交警吧，天天在路边抓人的车，唉，太别扭了。多多做点教育嘛",
            "来武汉教教吧，转弯后进哪条道一大半人都不知道",
            "右转能看的见后吗 这个规定就不对 非机动车过路口不该减速吗 自己要往上撞",
            "左转看右就是瞎扯，左转看左比看右重要得多"
        ]
    },
    {
        "视频 ID": "7619264361403534634",
        "标题": "使用军用 AI 打击我永兴岛？美国 AI 军火商露出獠牙（上）",
        "作者": "域与局",
        "视频链接": "https://v.douyin.com/cR3QGNmWJYc/",
        "视频时长": "06:16",
        "发布时间": "2026-03-20 19:20",
        "简介文案": "使用军用 AI 打击我永兴岛？美国 AI 军火商露出獠牙（上） #时代的荣耀 #全球创作者计划 #零基础看懂全球 #硬核深度计划",
        "作者声明": "内容来源于网络",
        "章节要点": """【总述】
美国 Palantir 科技公司是一家披着硅谷科创企业外衣的军工巨头，与美国极右翼势力、军工资本、犹太资本深度融合。其创始人彼得·蒂尔和亚历克斯·卡普利用技术优势与美国政府合作，为 CIA 开发"哥谭系统"，进军防务市场。Palantir 在俄乌冲突和加沙冲突中提供技术支持，股价暴涨，市值突破 3000 亿美元。

【章节 1 - 引言】00:01
美国 Palantir 公司发布模拟打击永兴岛的 AI 协同作战方案，展示其军用 AI 技术能力。Palantir 是一家披着硅谷科创企业外衣的军工巨头，与美国极右翼势力、军工资本、犹太资本深度融合。

【章节 2 - Palantir 的发展历程】02:25
Palantir 由彼得·蒂尔和亚历克斯·卡普创立，利用硅谷技术优势与美国政府合作，鼓吹"国家安全观"。

【章节 3 - Palantir 的商业模式】03:30
Palantir 为 CIA 开发"哥谭系统"，进军防务市场，但遭传统军工复合体抵制。

【章节 4 - Palantir 的政治影响力】04:46
彼得蒂尔支持特朗普，Palantir 成为美国军政复合体重要一环，政府订单不断。

【章节 5 - Palantir 的战争牟利】05:30
Palantir 在俄乌冲突和加沙冲突中提供技术支持，股价暴涨，市值突破 3000 亿美元。""",
        "弹幕/热评": [
            "这么厉害的军工科技企业好像只在前几年听到过一次，说美国欧洲的高科技公司已经投入鹅乌战争",
            "十年前还在清华听彼得蒂尔演讲",
            "中美战争，就是人工智能战争，建议国内人工智能企业全面融合国内军工企业",
            "那我们的战颅系统、千手观音算什么水平呢",
            "我就想知道，我国有没有类似的系统，另外有没有针对这种人进行暗中清除"
        ]
    }
]

# 创建 Excel 工作簿
wb = Workbook()

# Sheet 1: 文案汇总
ws_summary = wb.active
ws_summary.title = "文案汇总"

# 定义样式
title_font = Font(bold=True, size=14, color="000000")
header_font = Font(bold=True, size=11, color="FFFFFF")
normal_font = Font(size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
green_font = Font(color="008000")

# 表头
headers = ["序号", "视频标题", "作者", "视频时长", "简介文案", "章节要点", "热评摘录", "视频链接"]
for col, header in enumerate(headers, start=1):
    cell = ws_summary[f'{get_column_letter(col)}1']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 数据行
for idx, video in enumerate(videos, start=2):
    ws_summary[f'A{idx}'] = idx - 1
    ws_summary[f'B{idx}'] = video["标题"]
    ws_summary[f'C{idx}'] = video["作者"]
    ws_summary[f'D{idx}'] = video["视频时长"]
    ws_summary[f'E{idx}'] = video["简介文案"]
    ws_summary[f'F{idx}'] = video["章节要点"]
    ws_summary[f'G{idx}'] = " | ".join(video["弹幕/热评"][:3])  # 只显示前 3 条热评
    ws_summary[f'H{idx}'] = video["视频链接"]
    
    # 设置格式
    for col in range(1, 9):
        cell = ws_summary[f'{get_column_letter(col)}{idx}']
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 链接用绿色
    ws_summary[f'H{idx}'].font = green_font

# 设置列宽
ws_summary.column_dimensions['A'].width = 6
ws_summary.column_dimensions['B'].width = 40
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 10
ws_summary.column_dimensions['E'].width = 50
ws_summary.column_dimensions['F'].width = 50
ws_summary.column_dimensions['G'].width = 45
ws_summary.column_dimensions['H'].width = 50

# Sheet 2: 详细章节
ws_chapters = wb.create_sheet("详细章节")

ws_chapters['A1'] = "视频文案详细章节"
ws_chapters['A1'].font = title_font

row = 3
for idx, video in enumerate(videos, start=1):
    ws_chapters[f'A{row}'] = f"视频{idx}: {video['标题']}"
    ws_chapters[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws_chapters[f'A{row}'] = f"作者：{video['作者']}"
    ws_chapters[f'B{row}'] = f"时长：{video['视频时长']}"
    ws_chapters[f'C{row}'] = f"发布：{video['发布时间']}"
    row += 1
    
    ws_chapters[f'A{row}'] = "简介文案:"
    ws_chapters[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_chapters[f'A{row}'] = video["简介文案"]
    ws_chapters[f'A{row}'].alignment = Alignment(wrap_text=True)
    row += 2
    
    ws_chapters[f'A{row}'] = "章节要点:"
    ws_chapters[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_chapters[f'A{row}'] = video["章节要点"]
    ws_chapters[f'A{row}'].alignment = Alignment(wrap_text=True)
    row += 2
    
    ws_chapters[f'A{row}'] = "热门评论:"
    ws_chapters[f'A{row}'].font = Font(bold=True)
    row += 1
    for comment in video["弹幕/热评"]:
        ws_chapters[f'A{row}'] = f"• {comment}"
        ws_chapters[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 2
    ws_chapters[f'A{row}'] = "-" * 80
    row += 2

ws_chapters.column_dimensions['A'].width = 80
ws_chapters.column_dimensions['B'].width = 30
ws_chapters.column_dimensions['C'].width = 30

# Sheet 3: 热评汇总
ws_comments = wb.create_sheet("热评汇总")

ws_comments['A1'] = "视频热评汇总"
ws_comments['A1'].font = title_font

# 表头
comment_headers = ["序号", "视频标题", "评论内容", "评论类型"]
for col, header in enumerate(comment_headers, start=1):
    cell = ws_comments[f'{get_column_letter(col)}1']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 填充评论数据
row = 2
comment_idx = 1
for video in videos:
    for comment in video["弹幕/热评"]:
        ws_comments[f'A{row}'] = comment_idx
        ws_comments[f'B{row}'] = video["标题"][:30] + "..." if len(video["标题"]) > 30 else video["标题"]
        ws_comments[f'C{row}'] = comment
        ws_comments[f'D{row}'] = "弹幕" if len(comment) < 15 else "评论"
        row += 1
        comment_idx += 1

ws_comments.column_dimensions['A'].width = 6
ws_comments.column_dimensions['B'].width = 45
ws_comments.column_dimensions['C'].width = 60
ws_comments.column_dimensions['D'].width = 12

# 添加抓取时间
ws_comments[f'A{row+2}'] = "抓取时间"
ws_comments[f'B{row+2}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws_comments[f'A{row+3}'] = "备注"
ws_comments[f'B{row+3}'] = "抖音需要登录才能获取完整字幕，此处提取的是章节要点和页面可见文案"

# 保存文件
output_file = "/home/admin/openclaw/workspace/output/抖音视频文案提取汇总_3 个视频.xlsx"
wb.save(output_file)

print(f"✅ 文案 Excel 文件已创建：{output_file}")
print(f"\n📝 提取内容：")
for idx, video in enumerate(videos, start=1):
    print(f"\n视频{idx}: {video['标题']}")
    print(f"   作者：{video['作者']}")
    print(f"   时长：{video['视频时长']}")
    print(f"   章节数：{video['章节要点'].count('【')} 个")
    print(f"   热评数：{len(video['弹幕/热评'])} 条")
