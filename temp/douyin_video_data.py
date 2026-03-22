#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抖音视频数据提取并创建 Excel 报表
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 视频数据（从浏览器 snapshot 提取）
video_data = {
    "视频 ID": "7616988650777087295",
    "标题": "情绪价值拉满的语言技巧，推拉技巧",
    "作者": "灯哥说脱单",
    "作者主页": "https://www.douyin.com/user/MS4wLjABAAAAOAwjxUrLpBR7ZZMnEHH_c13TCmAeyH8J3DUTPeg8LHV8egfwPIPveH9mEJGHY_GO",
    "粉丝数": "9385",
    "总获赞": "1.7 万",
    "发布时间": "2026-03-14 19:27",
    "视频时长": "04:22",
    "点赞数": "2320",
    "评论数": "55",
    "收藏数": "1016",
    "分享数": "362",
    "话题标签": "#干货分享 #推拉技巧 #情绪价值 #恋爱心理学 #恋爱",
    "作者声明": "虚构演绎，仅供娱乐",
    "章节概要": """
1. 引言 (00:12)
2. 推拉技巧的定义 (00:21) - 推拉技巧是在相处中一进一退，让对方情绪波动，拉近关系
3. 先推后拉 (00:39) - 通过先拉低预期再突然拉高，形成情绪波动，效果更佳
4. 先拉后推 (01:30) - 用于维持自身框架，增加撩拨威力
5. 行为上的推拉 (02:23) - 推拉技巧可搭配行为使用，如身体前倾和后退，测试女生好感度
6. 结语 (04:11)
""",
    "视频链接": "https://v.douyin.com/KE4eK_vzT0Y/",
    "抓取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
}

# 创建 Excel 工作簿
wb = Workbook()
ws = wb.active
ws.title = "视频数据"

# 定义样式
title_font = Font(bold=True, size=14, color="000000")
header_font = Font(bold=True, size=11, color="FFFFFF")
normal_font = Font(size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 关键假设
blue_font = Font(color="0000FF")  # 硬编码输入
green_font = Font(color="008000")  # 链接

# 列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 50

# 添加数据
row = 1
ws['A1'] = "字段"
ws['B1'] = "内容"
ws['A1'].font = header_font
ws['B1'].font = header_font
ws['A1'].fill = header_fill
ws['B1'].fill = header_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

# 数据字段顺序（重要信息在前）
fields_order = [
    ("视频标题", "标题"),
    ("作者", "作者"),
    ("作者主页链接", "作者主页"),
    ("视频链接", "视频链接"),
    ("视频 ID", "视频 ID"),
    ("发布时间", "发布时间"),
    ("视频时长", "视频时长"),
    ("粉丝数", "粉丝数"),
    ("总获赞", "总获赞"),
    ("点赞数", "点赞数"),
    ("评论数", "评论数"),
    ("收藏数", "收藏数"),
    ("分享数", "分享数"),
    ("话题标签", "话题标签"),
    ("作者声明", "作者声明"),
    ("章节概要", "章节概要"),
    ("抓取时间", "抓取时间")
]

for i, (display_name, key) in enumerate(fields_order, start=2):
    ws[f'A{i}'] = display_name
    ws[f'B{i}'] = video_data[key]
    ws[f'A{i}'].font = normal_font
    ws[f'B{i}'].font = normal_font
    ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws[f'B{i}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 链接用绿色字体
    if "链接" in display_name or "URL" in display_name:
        ws[f'B{i}'].font = green_font

# 添加互动数据汇总（带公式）
summary_row = len(fields_order) + 3
ws[f'A{summary_row}'] = "互动数据汇总"
ws[f'A{summary_row}'].font = title_font
ws[f'A{summary_row}'].alignment = Alignment(horizontal='left')

summary_row += 1
ws[f'A{summary_row}'] = "总互动量"
ws[f'B{summary_row}'] = "=B10+B11+B12+B13"  # 点赞 + 评论 + 收藏 + 分享
ws[f'A{summary_row}'].font = Font(bold=True)
ws[f'B{summary_row}'].font = Font(bold=True, color="008000")  # 公式用绿色

# 添加互动率计算
summary_row += 1
ws[f'A{summary_row}'] = "互动率 (点赞/粉丝)"
ws[f'B{summary_row}'] = "=B10/B8*100"
ws[f'B{summary_row}'].number_format = "0.00%"
ws[f'A{summary_row}'].font = Font(bold=True)
ws[f'B{summary_row}'].font = Font(bold=True, color="008000")

# 添加推荐视频列表
rec_videos = [
    ("为什么说男人越放得开女人越爱你？", "拒绝废话（圈圈版）", "6585", "01:36"),
    ("真正的穷不是吃不起肉，而是家里有个永远在扫兴的人。", "心理叨叨兽", "8.0 万", "03:25"),
    ("停止讨好！三招让你从舔狗变海王！！！", "Paoen", "1.4 万", "04:35"),
    ("恋爱一定要遵循恋人优先原则", "悦叔 Yue", "16.5 万", "02:46"),
    ("分手后突然回头，是真的后悔还是权衡利弊后的选择？", "影鹿（叫我小鹿）", "8.0 万", "09:28"),
]

rec_row = summary_row + 3
ws[f'A{rec_row}'] = "推荐视频（同主题）"
ws[f'A{rec_row}'].font = title_font
ws[f'A{rec_row}'].alignment = Alignment(horizontal='left')

# 推荐视频表头
rec_row += 1
headers = ["视频标题", "作者", "播放量", "时长"]
for col, header in enumerate(headers, start=1):
    ws[f'{get_column_letter(col)}{rec_row}'] = header
    ws[f'{get_column_letter(col)}{rec_row}'].font = header_font
    ws[f'{get_column_letter(col)}{rec_row}'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws[f'{get_column_letter(col)}{rec_row}'].font = Font(bold=True, size=11, color="FFFFFF")

# 推荐视频数据
for i, (title, author, views, duration) in enumerate(rec_videos, start=1):
    row_num = rec_row + i
    ws[f'A{row_num}'] = title
    ws[f'B{row_num}'] = author
    ws[f'C{row_num}'] = views
    ws[f'D{row_num}'] = duration

# 设置推荐视频表格的列宽
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10

# 保存文件
output_file = "/home/admin/openclaw/workspace/output/抖音视频数据_灯哥说脱单.xlsx"
wb.save(output_file)

# 使用 recalc 重算公式
import subprocess
result = subprocess.run(
    ["python3", "/home/admin/openclaw/workspace/scripts/recalc.py", output_file],
    capture_output=True,
    text=True
)

print(f"✅ Excel 文件已创建：{output_file}")
print(f"📊 重算结果：{result.stdout}")
if result.stderr:
    print(f"⚠️ 警告：{result.stderr}")
