#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""创建测试用的 Excel 文件"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 创建工作簿
wb = Workbook()

# 重命名默认 sheet
wb.active.title = "汇总表"

# 创建明细表
detail_sheet = wb.create_sheet("明细表")

# 添加表头
headers = ['A 编号', 'B 姓名', 'C 部门', 'D 职位', 'E 入职日期', 'F 工资', 'G 绩效', 
           'H 奖金', 'I 备注', 'J 项目', 'K 状态', 'L 邮箱', 'M 电话', 'N 地址', 
           'O 紧急联系人', 'P 合同类型']
detail_sheet.append(headers)

# 添加测试数据
data = [
    [1001, '张三', '技术部', '工程师', '2023-01-15', 15000, 'A', 5000, '优秀员工', '项目 A', '在职', 'zhangsan@test.com', '138001', '北京', '李四', '正式'],
    [1002, '李四', '销售部', '销售', '2023-03-20', 12000, 'B', 3000, '', '项目 B', '在职', 'lisi@test.com', '138002', '上海', '王五', '正式'],
    [1003, '王五', '技术部', '架构师', '2022-06-10', 25000, 'A', 8000, '核心人员', '项目 A', '在职', 'wangwu@test.com', '138003', '深圳', '赵六', '正式'],
    [1004, '赵六', '人事部', 'HR', '2023-08-05', 10000, 'C', 2000, '', '项目 C', '在职', 'zhaoliu@test.com', '138004', '广州', '张三', '合同'],
    [1005, '钱七', '技术部', '工程师', '2024-01-10', 14000, 'B', 4000, '', '项目 B', '在职', 'qianqi@test.com', '138005', '杭州', '孙八', '正式'],
    [1006, '孙八', '财务部', '会计', '2023-05-15', 11000, 'A', 3500, '细心', '项目 D', '在职', 'sunba@test.com', '138006', '成都', '钱七', '正式'],
    [1007, '周九', '技术部', '测试', '2023-11-20', 13000, 'B', 3000, '', '项目 A', '试用期', 'zhoujiu@test.com', '138007', '武汉', '周十', '合同'],
    [1008, '吴十', '市场部', '市场', '2024-02-01', 12000, 'C', 2500, '', '项目 E', '在职', 'wushi@test.com', '138008', '南京', '吴十一', '正式'],
]

for row in data:
    detail_sheet.append(row)

# 隐藏一些行（第 4 行和第 6 行）
detail_sheet.row_dimensions[4].hidden = True
detail_sheet.row_dimensions[6].hidden = True

# 隐藏一些列（E 列和 H 列）
detail_sheet.column_dimensions['E'].hidden = True
detail_sheet.column_dimensions['H'].hidden = True

# 添加一些公式
detail_sheet['Q1'] = '总薪资'
detail_sheet['Q2'] = '=F2+H2'  # 工资 + 奖金
for i in range(3, 10):
    detail_sheet[f'Q{i}'] = f'=F{i}+H{i}'

# 创建一个隐藏的 sheet
hidden_sheet = wb.create_sheet("隐藏数据")
hidden_sheet.sheet_state = 'hidden'
hidden_sheet.append(['隐藏数据 1', '隐藏数据 2', '隐藏数据 3'])
hidden_sheet.append(['A', 'B', 'C'])
hidden_sheet.append(['X', 'Y', 'Z'])

# 保存文件
test_file = '/home/admin/openclaw/workspace/test_明细表.xlsx'
wb.save(test_file)

print(f"✅ 测试文件已创建：{test_file}")
print(f"📋 包含:")
print(f"   - 汇总表 (空)")
print(f"   - 明细表 (有数据，隐藏了第 4、6 行和 E、H 列)")
print(f"   - 隐藏数据 (隐藏的 sheet)")
print(f"   - 包含公式 (Q 列：工资 + 奖金)")
